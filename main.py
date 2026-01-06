# student_scheduler.py
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random
from collections import defaultdict
import copy
import os
import subprocess
import platform
import time


class StudentScheduler:
    def __init__(self, num_students, num_periods, num_choices, tolerance):
        self.num_students = num_students
        self.num_periods = num_periods  # 受講する講座数（例: 4）
        self.num_choices = num_choices  # 希望順位の数（例: 6、これが講座数）
        self.tolerance = tolerance
        self.students = []
        self.courses = []  # 全講座リスト
        self.input_file = "入力_生徒希望アンケート.xlsx"
        self.output_file = "出力_講座配置結果.xlsx"

    def create_input_template(self):
        """入力用のテンプレートExcelファイルを作成"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "アンケート入力"

        # スタイル定義
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        input_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        # ヘッダー行
        headers = ['生徒番号', '氏名'] + [f'第{i}希望' for i in range(1, self.num_choices + 1)]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        # データ入力行
        for row_idx in range(2, self.num_students + 2):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row_idx, col_idx, '')
                cell.border = border
                cell.fill = input_fill
                cell.alignment = left_align if col_idx == 2 else center_align

        # サンプルデータ（最初の3行）
        sample_data = [
            ['001', '山田太郎', 'プログラミング', '美術', '音楽', '体育', '英会話', '料理'],
            ['002', '佐藤花子', '美術', '音楽', '料理', 'プログラミング', '体育', '英会話'],
            ['003', '鈴木一郎', '体育', 'プログラミング', '英会話', '美術', '料理', '音楽'],
        ]

        for row_idx, data in enumerate(sample_data[:min(3, self.num_students)], 2):
            for col_idx, value in enumerate(data[:len(headers)], 1):
                ws.cell(row_idx, col_idx, value)

        # 列幅の調整
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        for col in range(3, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        # 行の高さ調整
        ws.row_dimensions[1].height = 25
        for row in range(2, self.num_students + 2):
            ws.row_dimensions[row].height = 22

        # 注意事項シート
        ws_info = wb.create_sheet("使い方", 0)
        ws_info.column_dimensions['A'].width = 80

        # 各時限の目標人数
        target_per_period_course = self.num_students * self.num_periods // (self.num_choices * self.num_periods)

        info_texts = [
            "【学生講座配置プログラム - 使い方】",
            "",
            "■ 入力手順",
            "1. 「アンケート入力」シートを開きます",
            "2. 黄色のセルに生徒情報と希望を入力します",
            "3. 入力が完了したら、ファイルを保存して閉じます",
            "4. プログラムが自動的に配置を計算します",
            "",
            "■ 入力項目",
            f"・生徒番号: 生徒を識別する番号（必須）",
            f"・氏名: 生徒の氏名（必須）",
            f"・第1希望〜第{self.num_choices}希望: 希望する講座名を入力",
            "",
            "■ 設定情報",
            f"・生徒数: {self.num_students}名",
            f"・講座数: {self.num_choices}講座（全講座が開講されます）",
            f"・受講数: 各生徒は{self.num_periods}講座を受講",
            f"・時限数: {self.num_periods}時限",
            f"・人数許容範囲: 目標 ±{self.tolerance}名",
            "",
            "■ 配置ルール",
            f"・各時限で全{self.num_choices}講座が開講されます",
            f"・各生徒は{self.num_choices}講座のうち{self.num_periods}講座を受講します",
            "・生徒によって受講する講座の組み合わせは異なります",
            "・できるだけ希望順位の高い講座が選ばれます",
            "・各講座の人数ができるだけ均等になるよう調整されます",
        ]

        for row, text in enumerate(info_texts, 1):
            cell = ws_info.cell(row, 1, text)
            if text.startswith("【"):
                cell.font = Font(bold=True, size=14, color='4472C4')
            elif text.startswith("■"):
                cell.font = Font(bold=True, size=11)
            else:
                cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws_info.row_dimensions[row].height = 20

        wb.save(self.input_file)
        print(f"\n✓ 入力用ファイルを作成しました: {self.input_file}")

    def open_excel_file(self, filename):
        """Excelファイルを開く"""
        abs_path = os.path.abspath(filename)

        try:
            if platform.system() == 'Windows':
                os.startfile(abs_path)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.call(['open', abs_path])
            else:  # Linux
                subprocess.call(['xdg-open', abs_path])
            return True
        except Exception as e:
            print(f"ファイルを開けませんでした: {e}")
            return False

    def wait_for_file_close(self, filename):
        """ファイルが閉じられるまで待機"""
        print(f"\n📝 {filename} を開いています...")
        print("データを入力して保存し、ファイルを閉じてください。")
        print("（ファイルを閉じると自動的に処理が続行されます）")

        # Excelがファイルを開くまで待機（最大30秒）
        print("\nExcelがファイルを開くのを待っています...", end="", flush=True)
        file_opened = False
        for _ in range(30):
            try:
                with open(filename, 'r+b'):
                    pass
                print(".", end="", flush=True)
                time.sleep(1)
            except (PermissionError, IOError):
                file_opened = True
                print("\n✓ Excelがファイルを開きました。編集してください。")
                break

        if not file_opened:
            print("\n")
            input("Excelでファイルを開いて編集し、保存して閉じたら Enter キーを押してください...")
            return

        # ファイルが閉じられるまで待機
        print("ファイルが閉じられるのを待っています...", end="", flush=True)
        while True:
            try:
                with open(filename, 'r+b'):
                    pass
                print("\n✓ ファイルが閉じられました。処理を続行します...")
                time.sleep(1)
                break
            except (PermissionError, IOError):
                print(".", end="", flush=True)
                time.sleep(2)

    def load_data(self):
        """Excelファイルからデータを読み込む"""
        wb = openpyxl.load_workbook(self.input_file)
        ws = wb['アンケート入力']

        all_courses = set()

        for row in ws.iter_rows(min_row=2, max_row=self.num_students + 1, values_only=True):
            if row[0] is None or str(row[0]).strip() == '':
                continue
            if row[1] is None or str(row[1]).strip() == '':
                continue

            preferences = []
            for i in range(2, 2 + self.num_choices):
                if i < len(row) and row[i] is not None and str(row[i]).strip() != '':
                    preferences.append(str(row[i]).strip())

            if preferences:
                student = {
                    'id': str(row[0]).strip(),
                    'name': str(row[1]).strip(),
                    'preferences': preferences
                }
                self.students.append(student)
                all_courses.update(preferences)

        wb.close()

        if len(self.students) == 0:
            raise ValueError("有効な生徒データが見つかりません")

        # 全講座をリスト化（希望順位のスコアでソート）
        course_scores = defaultdict(int)
        for student in self.students:
            for rank, course in enumerate(student['preferences']):
                course_scores[course] += (self.num_choices - rank)

        self.courses = sorted(all_courses, key=lambda c: course_scores[c], reverse=True)

        print(f"\n✓ 読み込み完了: {len(self.students)}名の生徒データ")
        print(f"✓ 講座数: {len(self.courses)}講座")

        print("\n【講座一覧】（人気順）")
        for i, course in enumerate(self.courses, 1):
            print(f"  {i}. {course} (スコア: {course_scores[course]})")

    def get_preference_rank(self, student, course):
        """生徒の希望順位を取得（1始まり、希望外は大きな値）"""
        if course in student['preferences']:
            return student['preferences'].index(course) + 1
        return 999

    def calculate_student_score(self, selected_courses, student):
        """生徒が選択した講座の希望順位合計を計算"""
        total = 0
        for course in selected_courses:
            rank = self.get_preference_rank(student, course)
            total += rank
        return total

    def calculate_fairness_score(self, course_selection, schedule):
        """
        配置のスコアを計算（小さいほど良い）
        - 希望順位の合計
        - 公平性ペナルティ
        - 人数バランスペナルティ
        """
        # 各生徒の満足度スコア
        student_scores = []
        for student in self.students:
            selected = course_selection.get(student['id'], set())
            score = self.calculate_student_score(selected, student)
            student_scores.append(score)

        total = sum(student_scores)

        # 公平性ペナルティ
        if student_scores:
            max_score = max(student_scores)
            min_score = min(student_scores)
            fairness_penalty = (max_score - min_score) * 10

            avg_score = total / len(student_scores)
            variance = sum((s - avg_score) ** 2 for s in student_scores) / len(student_scores)
            variance_penalty = variance * 3

            total += fairness_penalty + variance_penalty

        # 人数バランスペナルティ
        if schedule:
            target = len(self.students) / len(self.courses)
            for period in range(1, self.num_periods + 1):
                for course in self.courses:
                    count = sum(1 for s in self.students
                                if schedule.get(s['id'], {}).get(period) == course)
                    if abs(count - target) > self.tolerance:
                        total += (abs(count - target) - self.tolerance) ** 2 * 50

        return total

    def greedy_assign(self):
        """
        貪欲法による初期配置
        1. 各生徒に受講する講座を決定
        2. 時間割を作成
        """
        # ========== フェーズ1: 講座選択 ==========
        # course_selection[student_id] = set of courses
        course_selection = {student['id']: set() for student in self.students}

        # 各講座の受講者数カウント
        course_counts = {course: 0 for course in self.courses}
        target_per_course = len(self.students) * self.num_periods / len(self.courses)
        max_per_course = target_per_course + self.tolerance * len(self.students) / 10

        print(f"\n【講座選択フェーズ】")
        print(f"目標受講者数: 各講座 約{target_per_course:.1f}名")

        # 各生徒が num_periods 個の講座を選択
        for round_num in range(self.num_periods):
            # 各ラウンドで生徒をシャッフル
            students_shuffled = self.students.copy()
            random.shuffle(students_shuffled)

            for student in students_shuffled:
                best_course = None
                best_rank = 999

                # まだ選択していない講座から、希望順位が高いものを選ぶ
                for course in student['preferences']:
                    if course in course_selection[student['id']]:
                        continue  # 既に選択済み
                    if course_counts[course] >= max_per_course:
                        continue  # 定員オーバー
                    rank = self.get_preference_rank(student, course)
                    if rank < best_rank:
                        best_rank = rank
                        best_course = course

                # 希望にない講座も検討（人数が少ない講座）
                if best_course is None:
                    available = [c for c in self.courses
                                 if c not in course_selection[student['id']]]
                    if available:
                        best_course = min(available, key=lambda c: course_counts[c])

                if best_course:
                    course_selection[student['id']].add(best_course)
                    course_counts[best_course] += 1

        # ========== フェーズ2: 時間割作成 ==========
        # schedule[student_id][period] = course
        schedule = {student['id']: {} for student in self.students}

        print(f"\n【時間割作成フェーズ】")

        for period in range(1, self.num_periods + 1):
            # この時限の各講座の人数
            period_course_counts = {course: 0 for course in self.courses}

            # 生徒をシャッフル
            students_shuffled = self.students.copy()
            random.shuffle(students_shuffled)

            for student in students_shuffled:
                # この生徒が選択した講座のうち、まだ配置されていないもの
                selected = course_selection[student['id']]
                already_scheduled = set(schedule[student['id']].values())
                available = selected - already_scheduled

                if not available:
                    continue

                # 人数が少ない講座を優先
                best_course = min(available, key=lambda c: period_course_counts[c])

                schedule[student['id']][period] = best_course
                period_course_counts[best_course] += 1

        return course_selection, schedule

    def improve_assignment(self, course_selection, schedule, iterations=30000):
        """焼きなまし法で配置を改善"""
        best_selection = copy.deepcopy(course_selection)
        best_schedule = copy.deepcopy(schedule)
        best_score = self.calculate_fairness_score(best_selection, best_schedule)

        current_selection = copy.deepcopy(course_selection)
        current_schedule = copy.deepcopy(schedule)
        current_score = best_score

        temperature = 200.0
        cooling_rate = 0.9997

        print(f"\n配置を最適化中（初期スコア: {best_score:.1f}）", end="")

        for iteration in range(iterations):
            if iteration % 3000 == 0:
                print(".", end="", flush=True)

            # 操作を選択
            operation = random.choice(['swap_schedule', 'swap_course'])

            if operation == 'swap_schedule':
                # 同じ時限で2人の生徒の講座を交換
                period = random.randint(1, self.num_periods)
                students_in_period = [s for s in self.students
                                      if period in current_schedule[s['id']]]
                if len(students_in_period) < 2:
                    continue

                s1, s2 = random.sample(students_in_period, 2)
                c1 = current_schedule[s1['id']][period]
                c2 = current_schedule[s2['id']][period]

                if c1 == c2:
                    continue

                # 交換が有効か確認（お互いがその講座を選択しているか）
                if c2 not in current_selection[s1['id']]:
                    continue
                if c1 not in current_selection[s2['id']]:
                    continue

                # 交換後に重複がないか確認
                s1_others = set(current_schedule[s1['id']].values()) - {c1}
                s2_others = set(current_schedule[s2['id']].values()) - {c2}
                if c2 in s1_others or c1 in s2_others:
                    continue

                # 交換を試行
                new_schedule = copy.deepcopy(current_schedule)
                new_schedule[s1['id']][period] = c2
                new_schedule[s2['id']][period] = c1

                new_score = self.calculate_fairness_score(current_selection, new_schedule)

            else:  # swap_course
                # 2人の生徒の選択講座を1つずつ交換
                s1, s2 = random.sample(self.students, 2)

                # s1だけが持っている講座とs2だけが持っている講座を交換
                s1_only = current_selection[s1['id']] - current_selection[s2['id']]
                s2_only = current_selection[s2['id']] - current_selection[s1['id']]

                if not s1_only or not s2_only:
                    continue

                c1 = random.choice(list(s1_only))
                c2 = random.choice(list(s2_only))

                # 新しい選択を作成
                new_selection = copy.deepcopy(current_selection)
                new_selection[s1['id']] = (current_selection[s1['id']] - {c1}) | {c2}
                new_selection[s2['id']] = (current_selection[s2['id']] - {c2}) | {c1}

                # スケジュールも更新
                new_schedule = copy.deepcopy(current_schedule)

                # s1のスケジュールでc1をc2に置換
                for period, course in list(new_schedule[s1['id']].items()):
                    if course == c1:
                        new_schedule[s1['id']][period] = c2

                # s2のスケジュールでc2をc1に置換
                for period, course in list(new_schedule[s2['id']].items()):
                    if course == c2:
                        new_schedule[s2['id']][period] = c1

                new_score = self.calculate_fairness_score(new_selection, new_schedule)
                current_selection = new_selection

            # スコア改善または確率的に受け入れ
            delta = new_score - current_score
            if delta < 0 or random.random() < pow(2.718, -delta / temperature):
                if operation == 'swap_schedule':
                    current_schedule = new_schedule
                else:
                    current_selection = new_selection
                    current_schedule = new_schedule
                current_score = new_score

                if current_score < best_score:
                    best_selection = copy.deepcopy(current_selection)
                    best_schedule = copy.deepcopy(current_schedule)
                    best_score = current_score

            temperature *= cooling_rate

        print(f" 完了！（最終スコア: {best_score:.1f}）")
        return best_selection, best_schedule

    def save_results(self, course_selection, schedule):
        """結果をExcelファイルに保存"""
        wb = openpyxl.Workbook()

        # 共通スタイル定義
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        good_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        warning_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        bad_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        # ========== シート1: 生徒×時限配置結果 ==========
        ws_result = wb.active
        ws_result.title = "生徒別配置結果"

        # ヘッダー
        headers = ['生徒番号', '氏名'] + [f'{i}限' for i in range(1, self.num_periods + 1)]
        for col, header in enumerate(headers, 1):
            cell = ws_result.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align

        # 生徒データ（生徒番号順にソート）
        sorted_students = sorted(self.students, key=lambda s: s['id'])

        for row_idx, student in enumerate(sorted_students, 2):
            # 生徒番号
            cell = ws_result.cell(row_idx, 1, student['id'])
            cell.border = border
            cell.alignment = center_align

            # 氏名
            cell = ws_result.cell(row_idx, 2, student['name'])
            cell.border = border
            cell.alignment = left_align

            # 各時限の配置
            for period in range(1, self.num_periods + 1):
                course = schedule[student['id']].get(period, '')
                cell = ws_result.cell(row_idx, 2 + period, course)
                cell.border = border
                cell.alignment = left_align

                # 希望順位に応じて色分け
                rank = self.get_preference_rank(student, course)
                if rank <= 2:
                    cell.fill = good_fill
                elif rank <= 4:
                    cell.fill = warning_fill
                elif rank < 999:
                    cell.fill = bad_fill

        # 列幅調整
        ws_result.column_dimensions['A'].width = 12
        ws_result.column_dimensions['B'].width = 15
        for col in range(3, 3 + self.num_periods):
            ws_result.column_dimensions[get_column_letter(col)].width = 18

        # ========== シート2: 講座別名簿 ==========
        ws_roster = wb.create_sheet("講座別名簿")

        col_offset = 0
        for period in range(1, self.num_periods + 1):
            for course in self.courses:
                # 講座ヘッダー
                start_col = col_offset + 1
                cell = ws_roster.cell(1, start_col, f"【{period}限】{course}")
                cell.font = Font(bold=True, size=11, color='FFFFFF')
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
                ws_roster.merge_cells(start_row=1, start_column=start_col,
                                       end_row=1, end_column=start_col + 1)
                ws_roster.cell(1, start_col + 1).border = border

                # サブヘッダー
                ws_roster.cell(2, start_col, '生徒番号').fill = subheader_fill
                ws_roster.cell(2, start_col, '生徒番号').border = border
                ws_roster.cell(2, start_col, '生徒番号').alignment = center_align
                ws_roster.cell(2, start_col + 1, '氏名').fill = subheader_fill
                ws_roster.cell(2, start_col + 1, '氏名').border = border
                ws_roster.cell(2, start_col + 1, '氏名').alignment = center_align

                # この講座の生徒を生徒番号順でリスト
                course_students = []
                for student in self.students:
                    if schedule[student['id']].get(period) == course:
                        course_students.append(student)

                course_students.sort(key=lambda s: s['id'])

                for row_idx, student in enumerate(course_students, 3):
                    ws_roster.cell(row_idx, start_col, student['id']).border = border
                    ws_roster.cell(row_idx, start_col).alignment = center_align
                    ws_roster.cell(row_idx, start_col + 1, student['name']).border = border
                    ws_roster.cell(row_idx, start_col + 1).alignment = left_align

                # 人数表示
                count_row = max(len(course_students) + 3, 4)
                ws_roster.cell(count_row, start_col, f"計: {len(course_students)}名")
                ws_roster.cell(count_row, start_col).font = Font(bold=True)

                # 列幅調整
                ws_roster.column_dimensions[get_column_letter(start_col)].width = 10
                ws_roster.column_dimensions[get_column_letter(start_col + 1)].width = 12

                col_offset += 3  # 次の講座へ（1列空ける）

            col_offset += 1  # 次の時限へ（さらに1列空ける）

        # ========== シート3: 希望達成度 ==========
        ws_stats = wb.create_sheet("希望達成度")

        # ヘッダー
        stat_headers = ['生徒番号', '氏名', '満足度', '平均順位'] + \
                       [f'第{i}希望' for i in range(1, self.num_choices + 1)] + ['希望外']
        for col, header in enumerate(stat_headers, 1):
            cell = ws_stats.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align

        # 各生徒の統計を計算
        student_stats = []
        for student in sorted_students:
            rank_counts = defaultdict(int)
            total_rank = 0
            count = 0

            selected = course_selection.get(student['id'], set())
            for course in selected:
                rank = self.get_preference_rank(student, course)
                if rank < 999:
                    rank_counts[rank] += 1
                    total_rank += rank
                else:
                    rank_counts['希望外'] += 1
                    total_rank += self.num_choices + 1
                count += 1

            avg_rank = total_rank / count if count > 0 else 0
            # 満足度スコア: 100点満点
            max_possible = self.num_periods  # 全部第1希望
            min_possible = self.num_periods * (self.num_choices + 1)  # 全部希望外
            satisfaction = 100 * (min_possible - total_rank) / (min_possible - max_possible) if min_possible > max_possible else 100

            student_stats.append({
                'student': student,
                'satisfaction': satisfaction,
                'avg_rank': avg_rank,
                'rank_counts': rank_counts
            })

        # 満足度でソート（低い順）
        student_stats.sort(key=lambda x: x['satisfaction'])

        for row_idx, stat in enumerate(student_stats, 2):
            student = stat['student']

            cell = ws_stats.cell(row_idx, 1, student['id'])
            cell.border = border
            cell.alignment = center_align

            cell = ws_stats.cell(row_idx, 2, student['name'])
            cell.border = border
            cell.alignment = left_align

            cell = ws_stats.cell(row_idx, 3, round(stat['satisfaction'], 1))
            cell.border = border
            cell.alignment = center_align
            if stat['satisfaction'] >= 80:
                cell.fill = good_fill
            elif stat['satisfaction'] >= 60:
                cell.fill = warning_fill
            else:
                cell.fill = bad_fill

            cell = ws_stats.cell(row_idx, 4, round(stat['avg_rank'], 2))
            cell.border = border
            cell.alignment = center_align

            for rank in range(1, self.num_choices + 1):
                cell = ws_stats.cell(row_idx, 4 + rank, stat['rank_counts'].get(rank, 0))
                cell.border = border
                cell.alignment = center_align

            cell = ws_stats.cell(row_idx, 5 + self.num_choices, stat['rank_counts'].get('希望外', 0))
            cell.border = border
            cell.alignment = center_align

        # 統計サマリー
        summary_row = len(student_stats) + 3
        ws_stats.cell(summary_row, 1, '【統計】').font = Font(bold=True)

        satisfactions = [s['satisfaction'] for s in student_stats]
        avg_ranks = [s['avg_rank'] for s in student_stats]

        stats_info = [
            (summary_row + 1, '平均満足度', f"{sum(satisfactions)/len(satisfactions):.1f}点"),
            (summary_row + 2, '最低満足度', f"{min(satisfactions):.1f}点"),
            (summary_row + 3, '最高満足度', f"{max(satisfactions):.1f}点"),
            (summary_row + 4, '標準偏差', f"{(sum((s-sum(satisfactions)/len(satisfactions))**2 for s in satisfactions)/len(satisfactions))**0.5:.2f}"),
            (summary_row + 5, '平均希望順位', f"{sum(avg_ranks)/len(avg_ranks):.2f}"),
        ]

        for row, label, value in stats_info:
            ws_stats.cell(row, 1, label).font = Font(bold=True)
            ws_stats.cell(row, 2, value)

        # 列幅調整
        ws_stats.column_dimensions['A'].width = 12
        ws_stats.column_dimensions['B'].width = 15
        ws_stats.column_dimensions['C'].width = 12
        ws_stats.column_dimensions['D'].width = 12
        for col in range(5, 6 + self.num_choices):
            ws_stats.column_dimensions[get_column_letter(col)].width = 10

        wb.save(self.output_file)
        print(f"\n✓ 結果を保存しました: {self.output_file}")

    def print_summary(self, course_selection, schedule):
        """結果のサマリーを表示"""
        print("\n" + "=" * 70)
        print("配置結果サマリー")
        print("=" * 70)

        # 時限・講座別人数
        print("\n【時限・講座別人数】")
        target = len(self.students) / len(self.courses)
        for period in range(1, self.num_periods + 1):
            print(f"\n  {period}限:")
            for course in self.courses:
                count = sum(1 for s in self.students
                            if schedule[s['id']].get(period) == course)
                diff = count - target
                diff_str = f"+{diff:.0f}" if diff > 0 else f"{diff:.0f}"
                status = "✓" if abs(diff) <= self.tolerance else "!"
                print(f"    {course}: {count}名 ({diff_str}) {status}")

        # 希望達成状況
        print("\n【希望達成状況】")
        rank_counts = defaultdict(int)
        total_assignments = 0

        for student in self.students:
            selected = course_selection.get(student['id'], set())
            for course in selected:
                total_assignments += 1
                rank = self.get_preference_rank(student, course)
                if rank < 999:
                    rank_counts[rank] += 1
                else:
                    rank_counts['希望外'] += 1

        for rank in range(1, self.num_choices + 1):
            count = rank_counts.get(rank, 0)
            percentage = count / total_assignments * 100 if total_assignments > 0 else 0
            bar = "■" * int(percentage / 5)
            print(f"  第{rank}希望: {count:3d}件 ({percentage:5.1f}%) {bar}")

        hope_outside = rank_counts.get('希望外', 0)
        if hope_outside > 0:
            percentage = hope_outside / total_assignments * 100
            bar = "■" * int(percentage / 5)
            print(f"  希望外 : {hope_outside:3d}件 ({percentage:5.1f}%) {bar}")


def main():
    print("=" * 70)
    print("        学生講座配置プログラム")
    print("=" * 70)
    print()

    # 入力
    while True:
        try:
            num_students = int(input("生徒の人数を入力してください: "))
            if num_students > 0:
                break
            print("1以上の数値を入力してください。")
        except ValueError:
            print("数値を入力してください。")

    while True:
        try:
            num_choices = int(input("講座数（希望順位の数）を入力してください（例: 6）: "))
            if num_choices > 0:
                break
            print("1以上の数値を入力してください。")
        except ValueError:
            print("数値を入力してください。")

    while True:
        try:
            num_periods = int(input(f"受講する講座数を入力してください（1〜{num_choices}）: "))
            if 1 <= num_periods <= num_choices:
                break
            print(f"1〜{num_choices}の数値を入力してください。")
        except ValueError:
            print("数値を入力してください。")

    while True:
        try:
            tolerance = int(input("人数の許容範囲を入力してください（例: ±2なら 2）: "))
            if tolerance >= 0:
                break
            print("0以上の数値を入力してください。")
        except ValueError:
            print("数値を入力してください。")

    try:
        scheduler = StudentScheduler(num_students, num_periods, num_choices, tolerance)

        # テンプレート作成
        print("\n" + "=" * 70)
        print("ステップ1: 入力ファイルの準備")
        print("=" * 70)
        scheduler.create_input_template()

        # Excelファイルを開いて待機
        print("\n" + "=" * 70)
        print("ステップ2: データ入力")
        print("=" * 70)
        scheduler.open_excel_file(scheduler.input_file)
        scheduler.wait_for_file_close(scheduler.input_file)

        # データ読み込み
        print("\n" + "=" * 70)
        print("ステップ3: データ処理")
        print("=" * 70)
        scheduler.load_data()

        # 初期配置
        print("\n" + "=" * 70)
        print("ステップ4: 配置計算")
        print("=" * 70)
        course_selection, schedule = scheduler.greedy_assign()

        # 配置の改善
        course_selection, schedule = scheduler.improve_assignment(
            course_selection, schedule, iterations=30000
        )

        # 結果の表示
        scheduler.print_summary(course_selection, schedule)

        # 結果を保存
        print("\n" + "=" * 70)
        print("ステップ5: 結果の保存")
        print("=" * 70)
        scheduler.save_results(course_selection, schedule)

        # 結果ファイルを開く
        print("\n結果ファイルを開きます...")
        scheduler.open_excel_file(scheduler.output_file)

        print("\n" + "=" * 70)
        print("処理が完了しました！")
        print("=" * 70)

    except FileNotFoundError as e:
        print(f"\nエラー: ファイルが見つかりません - {e}")
    except ValueError as e:
        print(f"\nエラー: {e}")
    except Exception as e:
        print(f"\nエラーが発生しました: {e}")
        import traceback
        traceback.print_exc()

    input("\nEnterキーを押して終了...")


if __name__ == "__main__":
    main()
